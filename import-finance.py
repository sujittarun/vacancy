#!/usr/bin/env python3
"""Read Crescent Stays' revenue, nights and cost model into index.html.

    python3 import-finance.py "~/Downloads/Revenue.xlsx"           # report only
    python3 import-finance.py "~/Downloads/Revenue.xlsx" --write   # and splice

Two sources, one gate.

REVENUE AND NIGHTS come from the monthly sheets (Jul26, Jun26, ...), each of
which carries a per-bucket block: booked nights, a price total in thousands, and
a booking percentage. Eleven buckets across the five buildings the operator
still owns plus the two they sold.

THE COST MODEL comes from the "Costing" sheet: rent, salaries, electricity,
water, refills, maintenance, AMC, contingency, laundry, bonuses, replacements,
per building. It is a standing model rather than a dated actuals series, which
is what it says it is and what the app says it is.

THE GATE is the Availability tab. A month's revenue is only worth showing if the
nights behind it are the nights that were actually sold, so every (month,
building) figure is checked against the count of occupied cells in that month's
availability rows. This does three jobs at once:

  - it throws out the eight months before Apr 2025 with no comment needed: the
    Availability tab does not reach them, so they cannot be checked. They are
    also, as it happens, template rows nobody updated — three are byte-identical
    and several read over 100% occupancy — but the app does not have to know
    that. Unverifiable is enough.
  - it catches Nov 2025, where a Madhapur cell reads 481,933 nights and
    Rs 2.89bn.
  - and it finds the two REAL holes, which are the interesting part: TreeTops
    from May 2026 books ~120 more flat-nights a month than its revenue line
    covers (four long tenancies billed outside this sheet), and Apr 2026's
    Telecom Nagar line is 30 nights short.

A short month is not discarded. It is carried with the shortfall attached, so
the app can say "this revenue covers 147 of the 271 nights sold" instead of
quietly reporting a collapse in margin that is really a gap in bookkeeping.
"""
import sys, os, re, json, datetime, collections
import openpyxl

# Flags are not paths. `--write` as argv[1] was taken as the workbook and the
# run died on "openpyxl does not support  file format" — a real refusal to write
# that reads exactly like a corrupt file.
_args = [a for a in sys.argv[1:] if not a.startswith("-")]
XLSX = os.path.expanduser(_args[0]) if _args else "Revenue.xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))
CODES = ["TT", "MP", "TN", "BH"]          # the buildings with a financial line
OUT   = CODES + ["LP"]                    # + the one that has none, carried anyway
MONI = {m: i + 1 for i, m in enumerate(
    ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"])}
# the monthly sheets bucket by size within a building; the app thinks in buildings
BUCKET = {"Ttops 1bhk":"TT", "Treetops":"TT", "Mpur 1bhk":"MP", "Madhapur":"MP",
          "Tnagar Studio":"TN", "Tnagar 3bhk":"TN", "Tnagar 2bhk":"TN",
          "Bhills 3bhk":"BH", "Bhills 1bhk":"BH"}
COST_COL = {"Telecom Ngr":"TN", "Treetops":"TT", "Madhapur":"MP", "B.Hills":"BH"}
COST_LINES = ["Rent","Salaries","Electricity","Water","Refills","Maintenance",
              "AMC (Lift, Gen)","Contingency","Laundry","Apartment maintance",
              "Bonus / gifts","Replacements"]

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

# ── the availability tab, counted per month per building ─────────────────────
rows = list(wb["Availability"].iter_rows(values_only=True))
hdr = [str(h).strip().replace("\n", " ") if h else None for h in rows[1]]
def building(k, h):
    if h is None: return None
    hl = h.lower()
    if hl == "vacancy" or "vasavi" in hl or "mumba" in hl: return None   # sold
    # the two numbered blocks collide by header, so position decides — the same
    # rule import-book.py proves against the flat list
    return "TT" if k <= 12 else "MP" if k <= 22 else "TN" if k <= 32 else "BH" if k <= 40 else "LP"
cols = [(k, building(k, h)) for k, h in enumerate(hdr) if k > 0 and building(k, h)]

# A COLUMN IS A BEDROOM; A UNIT IS WHAT YOU LET. Lotus Pond's sheet keeps one
# column per bedroom — "101, 102, 103, Studio 104" under a merged "1st Floor"
# header — and 101/102/103 are one 3 BHK. Counted per column this building sold
# three nights every time it sold one, which is how a month came to report 387
# nights out of a capacity of 248.
#
# The book proves the grouping rather than assuming it: across 253 overlapping
# pairs of those rooms not one carries a different guest, and across the 88
# stays they describe not one has the three rooms on different dates. So a unit
# is sold when ANY of its columns is filled, and the trio counts once.
# LOTUS POND ONLY. The first pass matched any header shaped like 101/102/103
# and quietly merged TreeTops' rooms as well — TT's book fell from 271 nights to
# 182 and the gate then rejected eleven months it had been keeping. TreeTops is
# ten separate flats that happen to be numbered the same way; the merge is a
# fact about one building's floor plan, not about a numbering scheme.
def unit_of(k, b, h):
    if b != "LP" or h is None: return k
    m = re.fullmatch(r'(\d)0([123])(?:\.0)?', str(h).strip())
    return f"LP-{m.group(1)}" if m else k
units = {}
for k, b in cols:
    units.setdefault((b, unit_of(k, b, hdr[k])), []).append(k)
assert sum(1 for (b, _) in units if b == "TT") == 10, \
    f"TreeTops must stay ten separate flats, got {sum(1 for (b,_) in units if b=='TT')}"
assert sum(1 for (b, _) in units if b == "LP") == 8, \
    f"Lotus Pond should be 4 apartments and 4 studios, got {sum(1 for (b,_) in units if b=='LP')}"

sold = collections.Counter(); covered = set()
for r in rows[2:]:
    d = r[0]
    if isinstance(d, datetime.datetime): d = d.date()
    if not isinstance(d, datetime.date): continue
    covered.add((d.year, d.month))
    for (b, _), ks in units.items():
        if any(r[k] is not None and str(r[k]).strip() for k in ks):
            sold[(d.year, d.month, b)] += 1

# ── the monthly sheets ───────────────────────────────────────────────────────
series = {}
for name in wb.sheetnames:
    m = re.fullmatch(r'([A-Z][a-z]{2})(\d{2})', name)
    if not m or m.group(1) not in MONI: continue
    y, mo = 2000 + int(m.group(2)), MONI[m.group(1)]
    body = list(wb[name].iter_rows(max_row=60, max_col=6, values_only=True))
    h = next((i for i, r in enumerate(body)
              if any(isinstance(c, str) and c.strip() == "Booked" for c in r)), None)
    if h is None: continue
    got = {}
    for r in body[h + 1:h + 16]:
        k = r[0].strip() if isinstance(r[0], str) else None
        if k == "Total": break
        if k in BUCKET and isinstance(r[2], (int, float)) and isinstance(r[3], (int, float)):
            d = got.setdefault(BUCKET[k], [0, 0.0])
            d[0] += r[2]; d[1] += r[3]                 # nights, thousands of rupees
    if got: series[f"{y}-{mo:02d}"] = {k: [int(v[0]), round(v[1] * 1000)] for k, v in got.items()}

# ── the gate ─────────────────────────────────────────────────────────────────
today = datetime.date.today()
kept, dropped, short = [], [], []
for key in sorted(series):
    y, mo = map(int, key.split("-"))
    if (y, mo) >= (today.year, today.month):
        dropped.append((key, "the month is not over")); continue
    if (y, mo) not in covered:
        dropped.append((key, "no availability rows to check it against")); continue
    row, bad = {}, None
    for c in OUT:
        n, rev = series[key].get(c, [0, 0])
        book = sold.get((y, mo, c), 0)
        if not book: continue
        if not n:
            # counted in the availability book, priced in no sheet at all. The
            # nights are still real and still carried: a building the P&L has
            # never heard of is the finding, not a row to leave out.
            row[c] = [None, None, book]; continue
        if n > book * 1.02:                 # the sheet claims more nights than were sold
            bad = f"{c} bills {n} nights against {book} in the book"; break
        row[c] = [n, rev, book]
        if book > n * 1.02: short.append((key, c, n, book))
    if bad: dropped.append((key, bad)); continue
    if any(v[0] is not None for v in row.values()): kept.append((key, row))

# ── the cost model ───────────────────────────────────────────────────────────
crows = list(wb["Costing"].iter_rows(max_row=20, max_col=9, values_only=True))
chdr = [str(c).strip() if c else None for c in crows[0]]
cidx = {COST_COL[h]: i for i, h in enumerate(chdr) if h in COST_COL}
cost = {c: {} for c in cidx}
stated = {}
for r in crows:
    k = str(r[0]).strip() if r[0] else ""
    if k in COST_LINES:
        for c, i in cidx.items():
            if isinstance(r[i], (int, float)) and r[i]: cost[c][k] = int(r[i])
    if k == "Total Cost":
        for c, i in cidx.items(): stated[c] = int(r[i])

print(f"\nmonths kept {len(kept)}   dropped {len(dropped)}")
for k, why in dropped: print(f"   drop {k}  — {why}")
print(f"\ncost model  (line items vs the sheet's own Total Cost)")
allgood = True
for c in CODES:
    s = sum(cost[c].values())
    ok = s == stated.get(c)
    allgood &= ok
    print(f"   {c}  {s:>9,}  sheet {stated.get(c,0):>9,}  {'match' if ok else 'MISMATCH'}")
print(f"\nrevenue lines short of the book ({len(short)}):")
for k, c, n, book in short: print(f"   {k} {c}  covers {n} of {book} nights sold  (-{book-n})")
if not kept or not allgood:
    print("\nFAIL — nothing written"); sys.exit(1)
print("\nPASS — every kept month reconciles against the availability book")

if "--write" not in sys.argv:
    print("(dry run — pass --write to splice into index.html)"); sys.exit(0)

def cell(v): return json.dumps(v, ensure_ascii=False) if isinstance(v, str) else str(v)
def bcell(v):
    if v is None: return "null"
    return "[" + ",".join("null" if x is None else str(x) for x in v) + "]"
fin = ",\n".join("[" + cell(k) + "," + ",".join(bcell(r.get(c)) for c in OUT) + "]"
                 for k, r in kept)
drop = ",\n".join("[" + cell(k) + "," + cell(why) + "]" for k, why in dropped
                  if "not over" not in why)
cm = ",\n".join('  ' + c + ": {" + ", ".join(f'{json.dumps(k)}: {v}' for k, v in cost[c].items()) + "}"
                for c in CODES)
p = os.path.join(HERE, "index.html")
h = open(p).read()
for name, body, br in (("FIN", fin, "["), ("FIN_DROP", drop, "["), ("FIN_COST", cm, "{")):
    open_ = "const " + name + " = " + br + "\n"
    a = h.index(open_) + len(open_)
    # from a-1, not a: an EMPTY block is "[\n];" and its newline was just eaten
    # by the opening anchor, so a search from a walks straight past the closer
    # and lands on the NEXT block's — silently swallowing the declaration
    # between them. Cost an hour the first time.
    h = h[:a] + body + h[h.index("\n];" if br == "[" else "\n};", a - 1):]
open(p, "w").write(h)
print(f"written to {p}")
